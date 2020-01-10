import pandas as pd
import numpy as np
from openpyxl import load_workbook
from time import sleep
import os
from requests import Session
import json
import io
#Escrito por Bruno Gonzaga, Departamento de Análisis de Coyuntura
################################################################################################
# 1. Parametros de las busquedas

base_url = 'https://comtrade.un.org/api/get?'
frequency = 'A'     # anual
reporter = 'all'    # todos los países
partner = '0'         # 0 = world
tradeflow = '2'     # 2 = exports, 1 = imports
period = '2018'     # pueden listarse varios años con comas: 2016,2017,2018 (max. 5)
productdict = {'asp_f': '070920','asp_p':'200560','arti':'200599','blueb':'081040',
              'avoc':'080440','mang':'080450','grap':'080610'}
fmt = 'csv'         # formato del output
head = 'H'          # H = títulos son "human-readable" (tildes, comas, mayusculas, etc)

numero_consultas = len(productdict.keys())
registro_local = 0
path = r"E:\Users\2371\Desktop\Comtrade\comtrade-2018.xlsx"

################################################################################################
# 2. Bucle para cada request: URL + parametros

for value1 in productdict:
    registro_local = registro_local + 1
    parameters = {
            'ps': period,
            'freq': frequency,
            'r': reporter,
            'p': partner,
            'cc': productdict[value1],
            'rg': tradeflow,
            'px': 'HS',      # Harmonized System ('as reported')
            'type': 'C',     # C = commodities, S = services
            'fmt': fmt,      # formato del output
            'head': head     # human-readable ('H') o machine readable ('M')
        }

    string = ''
    for key,value in parameters.items():
        temp = '='.join([key, value])
        string = string + '&' + temp
    print(string)

    url = base_url + string[1:]
    print(url)
    print('Procesando registro '+ str(registro_local) + ' de ' + str(numero_consultas) + '...')

    ############################################################################################
    # 2.1 Identificacion del IP
    
    class config(object):

        proxy_url  = "http://2371:ene2020.@bcrproxy.bcrp.gob:8080"
        proxies = {'ftp': proxy_url,
                   'http': proxy_url,
                   'https': proxy_url}
        url_download = url

    session = Session()
    
    ############################################################################################
    # 2.2 Request
    
    r = session.get(proxies=config.proxies,
                    url=config.url_download)
    
    ############################################################################################
    # 2.3 Transformacion del dataframe
    
    df = pd.read_csv(io.StringIO(r.text))
    df.replace(['Commodity Code'], productdict[value1])
    dataframe = df[['Classification','Year','Reporter',
                 'Reporter ISO','Commodity Code','Commodity',
                        'Netweight (kg)','Trade Value (US$)']]
    dataframe = dataframe.sort_values(by='Trade Value (US$)',ascending = False)
    
    lista_temp = []
    for i in range(1,len(dataframe.index)+1):
        lista_temp.append(i)
    
    dataframe.insert(0, "Ranking", lista_temp, True)
    dataframe.insert(9, "Ranking", lista_temp, True)
    
    ############################################################################################
    # 2.4 Carga del archivo total, graba el dataframe en la pestaña correspondiente
    
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    
    try:
        del writer.book[value1]
    except:
        pass
        
    dataframe.to_excel(writer, sheet_name = value1, index = False)
    writer.save()
    writer.close()
    print('Espere un segundo...')
    
    ############################################################################################
    # 2.5 Descanso de 1s para no exceder límite de velocidad de descarga del API (1 req / seg)
    sleep(1)
    # FALTA: funcion para contar el numero de requests hechos en la ultima hora

print('Actualización completa')

#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%

# Guardar listado de codigos de productos
# NO ES NECESARIO EJECUTARLO SIEMPRE
url = 'https://comtrade.un.org/data/cache/classificationHS.json'
json_dict = session.get(proxies=config.proxies,url = url).json()
df = pd.DataFrame.from_dict(json_dict['results'])
df = df.set_index('id')
df.drop(['ALL', 'TOTAL', 'AG2', 'AG4', 'AG6'], inplace=True)
df.text = df.text.apply(lambda x: ' - '.join(x.split(' - ')[1:]))
df.to_excel('classificationHS.xlsx')

#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%

# Busqueda: MANDARINAS

base = pd.read_excel('classificationHS.xlsx', index_col='id')

var11 = base.text.str.contains('tangerines', case=True, flags=0, regex=True)
var12 = base.text.str.contains('mandarins', case=True, flags=0, regex=True)
var13 = base.text.str.contains('clementines', case=True, flags=0, regex=True)
var14 = base.text.str.contains('tangelos', case=True, flags=0, regex=True)
var2 = base.index.to_series().apply(lambda digit: len(digit) == 6)
var = (var11 | var12 | var13 | var14) & var2
mandarinas = base.text[var].to_dict()
print(mandarinas)

# Busqueda: POTA

base = pd.read_excel('classificationHS.xlsx', index_col='id')

var1 = base.text.str.contains('squid', case=True, flags=0, regex=True)
var2 = base.index.to_series().apply(lambda digit: len(digit) == 6)
var = (var1) & var2
pota = base.text[var].to_dict()
print(pota)

#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%%#%

# CODIGO PARA EXTRAER DATOS DE MANDARINAS

base_url = 'https://comtrade.un.org/api/get?'
frequency = 'A'     # anual
reporter = 'all'    # todos los países
partner = 0         # 0 = world
tradeflow = '2'     # 2 = exports, 1 = imports
period = '2018'     # pueden listarse varios años con comas: 2016,2017,2018 (max. 5)
fmt = 'csv'         # formato del output
head = 'H'          # H = títulos son "human-readable" (tildes, comas, mayusculas, etc)

numero_consultas = len(mandarinas.keys())
registro_local = 0
path1 = r"E:\Users\2371\Desktop\Comtrade\comtrade-2018.xlsx"

local_temp = 0;
df_mandarinas = pd.DataFrame()

for key1 in mandarinas:
    registro_local = registro_local + 1
    parameters = {
            'ps': period,
            'freq': frequency,
            'r': reporter,
            'p': partner,
            'cc': key1,
            'rg': tradeflow,
            'px': 'HS',      # Harmonized System ('as reported')
            'type': 'C',     # C = commodities, S = services
            'fmt': fmt,      # formato del output
            'head': head     # human-readable ('H') o machine readable ('M')
        }

    
    parameters2 = {}

    for key,value in parameters.items():
        value_string = str(value) if not isinstance(value, list) else ','.join(map(str, value))
        parameters2[key] = value_string

    string = ''
    for key,value in parameters2.items():
        temp = '='.join([key, value])
        string = string + '&' + temp
    print(string)

    url = base_url + string[1:]
    print(url)
    print('Procesando registro '+ str(registro_local) + ' de ' + str(numero_consultas) + '...')

    
    class config(object):

        proxy_url  = "http://2371:ene2020.@bcrproxy.bcrp.gob:8080"
        proxies = {'ftp': proxy_url,
                   'http': proxy_url,
                   'https': proxy_url}
        url_download = url

    session = Session()
    
    
    r = session.get(proxies=config.proxies,
                    url=config.url_download)
    
    local_temp = local_temp + 1

    df = pd.read_csv(io.StringIO(r.text))
    df.replace(['Commodity Code'], key1)
    df.sort_values(by=['Trade Value (US$)'])
    dataframe = df[['Reporter','Reporter ISO','Netweight (kg)','Trade Value (US$)']]
    dataframe = dataframe.sort_values(by='Trade Value (US$)',ascending = False)
    locals()['dataframe' + str(local_temp)] = dataframe
    df_mandarinas = pd.concat([df_mandarinas, dataframe],sort=True).groupby(['Reporter', 'Reporter ISO']).sum().reset_index()

df_mandarinas.insert(0,"Classification","Varios",allow_duplicates = False)
df_mandarinas.insert(1,"Year",2018,allow_duplicates = False)
df_mandarinas.insert(4,"Commodity Code","mandarinas",allow_duplicates = False)
df_mandarinas.insert(5,"Commodity","mandarinas",allow_duplicates = False)
df_mandarinas = df_mandarinas.sort_values(by='Trade Value (US$)',ascending = False)

lista_temp = []
for i in range(1,len(df_mandarinas.index)+1):
    lista_temp.append(i)
    
df_mandarinas.insert(0, "Ranking", lista_temp, True)
df_mandarinas.insert(9, "Ranking", lista_temp, True)

book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book

try:
    del writer.book["mand"]
except:
    pass

df_mandarinas.to_excel(writer, sheet_name = "mand", index = False)
writer.save()
writer.close()
print('Terminado')
